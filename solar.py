from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles import Font


inversor = {2000:330 , 
        2500: 360, 
        3000: 400,
        3300: 420,
        4200:520,
        5000:570,
        6000:610,
        7000:860,
        8000:880,
        9000:1000,
        10000:1020,
        }





def calcular_paneles(kw_bimestral):
    total_paneles = 0

    total_paneles = (kw_bimestral/0.5)

    total_paneles = (total_paneles /0.6)

    total_paneles = total_paneles/550

    total_paneles = round(total_paneles)

    costo_paneles = total_paneles*145*20 # costo por panel solar

    costo_paneles += (50*20)  # Complementos, siempre 1 gabinete?

    costo_paneles += (650*total_paneles) # aceros alcalde

    costo_paneles += (1000*total_paneles) # mano de obra

    costo_paneles += 1000+1000+500+1500 #Gastos de mano de obra, Cable, Tornillos,Complementos

    return total_paneles,costo_paneles


def calcular_inversor(total_paneles):

    capacidad_instalar = total_paneles*550

    print(capacidad_instalar)

    modelo_invesor =  min(inversor, key=lambda x: abs(x - capacidad_instalar))

    print(modelo_invesor)

    costo_invesor = inversor[modelo_invesor]*20

    return (capacidad_instalar/1000),modelo_invesor,costo_invesor

def basic_info(worksheet,info_basica):

    worksheet.merge_cells('B1:E1')
    worksheet.merge_cells('F1:G1')
    worksheet.merge_cells('B2:E2')
    worksheet.merge_cells('A3:B3')
    worksheet.merge_cells('F3:G3')
    worksheet.merge_cells('B5:D5')
    worksheet.merge_cells('B6:D6')
    worksheet.merge_cells('B7:D7')
    worksheet.merge_cells('B8:D8')
    worksheet.merge_cells('B9:D9')

    worksheet['A1'] = 'Nombre:'
    worksheet['B1'] = info_basica[0]
    worksheet['F1'] = 'Produccion Promedio: '
    worksheet['H1'] = 'TESTEO'

    worksheet['A2'] = 'Direccion: '
    worksheet['B2'] = info_basica[1]
    worksheet['F2'] = 'Tarifa: '
    worksheet['G2'] = 'TESTEO'

    worksheet['A3'] = 'No. de Cotizacion: '
    worksheet['C3'] = info_basica[2]
    worksheet['F3'] = 'Numero de Servicio: '
    worksheet['H3'] = 'TESTEO'

    worksheet['A5'] = 'Partida'
    worksheet['B5'] = 'Descripcion'
    worksheet['E5'] = 'Cantidad'
    worksheet['F5'] = 'Precio'

    bold_font = Font(bold=True)
    worksheet['A1'].font = bold_font
    worksheet['A2'].font = bold_font
    worksheet['A3'].font = bold_font
    worksheet['F1'].font = bold_font
    worksheet['F2'].font = bold_font
    worksheet['F3'].font = bold_font

    worksheet['A5'].font = bold_font
    worksheet['B5'].font = bold_font
    worksheet['E5'].font = bold_font
    worksheet['F5'].font = bold_font

    center_alignment = Alignment(horizontal='center', vertical='center')
    worksheet['A5'].alignment = center_alignment
    worksheet['B5'].alignment = center_alignment
    worksheet['E5'].alignment = center_alignment
    worksheet['F5'].alignment = center_alignment

def cotizacion(worksheet,paneles_info,inversor_info):

    worksheet['A6'] = 1
    worksheet ['B6'] = "PANEL SOLAR LONGI 550W monocristalino 144 celulas (6x24)"
    worksheet ['E6'] = paneles_info[0]
    worksheet ['F6'] = paneles_info[1]

    worksheet['A7'] = 2
    worksheet ['B7'] = "INVERSOR GROWATT "+str(inversor_info[1])+" TLX"
    worksheet ['E7'] = 1
    worksheet ['F7'] = inversor_info[2]






if __name__=="__main__":

    workbook = Workbook()
    worksheet = workbook.active
    paneles_info =  calcular_paneles(767)

    inversor_info = calcular_inversor(paneles_info[0])

    # print(paneles_info)
    # print(inversor_info)
    info_basica = ['Omar Guillermo Amezquita Alaniz','Koricancha 322','JKL897']
    basic_info(worksheet,info_basica)
    cotizacion(worksheet,paneles_info,inversor_info)
    workbook.save('test1.xlsx')

 
 





